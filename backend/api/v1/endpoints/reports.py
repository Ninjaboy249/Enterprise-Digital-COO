"""
Reports API Endpoints
- Upload Power BI (.pbix metadata JSON export) or raw JSON reports
- Upload Excel (.xlsx / .xls) — parsed, converted to Power BI-style structure, AI summarised
- Parse and extract structured data
- Summarise with OpenAI GPT
- Chat Q&A against uploaded report data
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
from typing import Dict, Any, Optional, List
from datetime import datetime
import json
import zipfile
import io
import re

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from openai import OpenAI
from config import settings

router = APIRouter()

# ─────────────────────────────────────────────────────────────────
# OpenAI client — created lazily so missing key doesn't crash boot
# ─────────────────────────────────────────────────────────────────

def _openai_client() -> OpenAI:
    key = settings.OPENAI_API_KEY
    if not key or key == "your-openai-api-key":
        raise HTTPException(
            status_code=503,
            detail="OpenAI API key is not configured. Set OPENAI_API_KEY in your .env file."
        )
    return OpenAI(api_key=key)


def _chat(client: OpenAI, system: str, user: str, max_tokens: int = 1200) -> str:
    """Single-turn GPT call, returns reply text."""
    model = settings.OPENAI_MODEL or "gpt-4o-mini"
    # Fall back gracefully for older keys
    try:
        rsp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user",   "content": user},
            ],
            max_tokens=max_tokens,
            temperature=0.4,
        )
        return rsp.choices[0].message.content.strip()
    except Exception as e:
        # Try cheaper model as fallback
        try:
            rsp = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user",   "content": user},
                ],
                max_tokens=max_tokens,
                temperature=0.4,
            )
            return rsp.choices[0].message.content.strip()
        except Exception as e2:
            raise HTTPException(status_code=502, detail=f"OpenAI call failed: {e2}")


# ─────────────────────────────────────────────────────────────────
# Parsers
# ─────────────────────────────────────────────────────────────────

def _parse_json_report(raw: bytes) -> Dict[str, Any]:
    """Parse a raw JSON file and return a normalised report dict."""
    try:
        data = json.loads(raw.decode("utf-8"))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON — cannot parse file.")

    # Flatten nested dict into flat key:value pairs for display & summarisation
    def _flatten(obj, prefix=""):
        items = {}
        if isinstance(obj, dict):
            for k, v in obj.items():
                items.update(_flatten(v, f"{prefix}{k}."))
        elif isinstance(obj, list):
            items[prefix.rstrip(".")] = f"[list of {len(obj)} items]"
            # Also try to summarise the first few entries
            for i, v in enumerate(obj[:5]):
                items.update(_flatten(v, f"{prefix}{i}."))
        else:
            items[prefix.rstrip(".")] = obj
        return items

    flat = _flatten(data)
    return {
        "type": "json",
        "source_data": data,
        "flat_preview": dict(list(flat.items())[:120]),   # cap for display
        "key_count": len(flat),
        "top_level_keys": list(data.keys()) if isinstance(data, dict) else [],
    }


def _parse_pbix(raw: bytes, filename: str) -> Dict[str, Any]:
    """
    A .pbix file is a ZIP archive.  Extract readable text artefacts:
      - DataModelSchema  (JSON — measures, tables, columns)
      - Report/Layout    (JSON — visuals, pages)
      - Connections      (XML/JSON — data sources)
    Returns a structured dict with what was found.
    """
    found: Dict[str, Any] = {"type": "pbix", "files_found": [], "sections": {}}
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            names = zf.namelist()
            found["files_found"] = names

            for name in names:
                lower = name.lower()

                # DataModelSchema — tables, measures, columns
                if "datamodelschema" in lower:
                    try:
                        content = json.loads(zf.read(name).decode("utf-8", errors="replace"))
                        model = content.get("model", content)
                        tables = model.get("tables", [])
                        found["sections"]["data_model"] = {
                            "tables": [
                                {
                                    "name": t.get("name", ""),
                                    "columns": [c.get("name", "") for c in t.get("columns", [])],
                                    "measures": [m.get("name", "") for m in t.get("measures", [])],
                                }
                                for t in tables
                            ]
                        }
                    except Exception:
                        found["sections"]["data_model"] = {"raw_snippet": str(zf.read(name)[:500])}

                # Report Layout — pages, visuals
                elif "report/layout" in lower or name == "Report/Layout":
                    try:
                        layout = json.loads(zf.read(name).decode("utf-8", errors="replace"))
                        pages = layout.get("sections", [])
                        found["sections"]["report_layout"] = {
                            "pages": [
                                {
                                    "name": p.get("displayName", p.get("name", "")),
                                    "visual_count": len(p.get("visualContainers", [])),
                                    "visuals": [
                                        _safe_visual_type(vc)
                                        for vc in p.get("visualContainers", [])[:10]
                                    ],
                                }
                                for p in pages
                            ]
                        }
                    except Exception:
                        found["sections"]["report_layout"] = {"raw_snippet": str(zf.read(name)[:500])}

                # Connections
                elif "connections" in lower:
                    try:
                        raw_txt = zf.read(name).decode("utf-8", errors="replace")
                        try:
                            conn = json.loads(raw_txt)
                        except Exception:
                            conn = {"raw": raw_txt[:800]}
                        found["sections"]["connections"] = conn
                    except Exception:
                        pass

    except zipfile.BadZipFile:
        raise HTTPException(
            status_code=400,
            detail=(
                "This .pbix file cannot be read as a ZIP archive. "
                "Power BI Desktop encrypts files saved from cloud workspaces. "
                "Please export a JSON/CSV from Power BI instead, or use a locally-saved .pbix."
            )
        )

    # If nothing useful was found, surface the file list
    if not found["sections"]:
        found["sections"]["notice"] = {
            "message": "No readable schema or layout found inside the .pbix.",
            "files": names,
        }

    return found


def _safe_visual_type(vc: dict) -> str:
    """Extract a visual type label from a Layout visualContainer entry."""
    try:
        cfg = json.loads(vc.get("config", "{}"))
        return cfg.get("singleVisual", {}).get("visualType", "unknown")
    except Exception:
        return "unknown"


def _parse_excel(raw: bytes, filename: str) -> Dict[str, Any]:
    """
    Parse an Excel workbook (.xlsx / .xls) into a Power BI-style structure.
    Extracts every sheet as a named table with columns, row count, sample rows,
    numeric summaries (min/max/mean), and detected KPI columns.
    """
    if not _OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="Excel support requires openpyxl. Add 'openpyxl' to your dependencies."
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {e}")

    sheets_info = []
    all_tables  = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # First non-empty row = headers
        headers = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(rows[0])]
        data_rows = rows[1:]

        # Cap at 500 rows for processing
        data_rows = data_rows[:500]

        # Build column-wise data
        col_data: Dict[str, list] = {h: [] for h in headers}
        for row in data_rows:
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                col_data[h].append(val)

        # Numeric summary per column
        col_summaries = {}
        kpi_cols = []
        for h, vals in col_data.items():
            nums = [v for v in vals if isinstance(v, (int, float)) and v is not None]
            if nums:
                col_summaries[h] = {
                    "type":  "numeric",
                    "count": len(nums),
                    "min":   round(min(nums), 4),
                    "max":   round(max(nums), 4),
                    "mean":  round(sum(nums) / len(nums), 4),
                    "sum":   round(sum(nums), 4),
                }
                kpi_cols.append(h)
            else:
                unique = list({str(v) for v in vals if v is not None})[:10]
                col_summaries[h] = {
                    "type":    "categorical",
                    "count":   len([v for v in vals if v is not None]),
                    "samples": unique,
                }

        # Sample rows (first 5, serialised as dicts)
        sample_rows = [
            {h: (str(row[i]) if i < len(row) else None) for i, h in enumerate(headers)}
            for row in data_rows[:5]
        ]

        sheet_info = {
            "sheet_name":    sheet_name,
            "row_count":     len(data_rows),
            "column_count":  len(headers),
            "columns":       headers,
            "kpi_columns":   kpi_cols,
            "column_stats":  col_summaries,
            "sample_rows":   sample_rows,
        }
        sheets_info.append(sheet_info)
        all_tables.append({"name": sheet_name, "columns": headers,
                           "measures": kpi_cols})

    wb.close()

    return {
        "type":     "excel",
        "filename": filename,
        # Power BI-compatible structure so existing chat/summary code works
        "sections": {
            "data_model": {"tables": all_tables},
            "sheets":     {s["sheet_name"]: s for s in sheets_info},
        },
        "sheets":       sheets_info,
        "sheet_count":  len(sheets_info),
        "files_found":  wb.sheetnames,
    }


def _report_to_text(parsed: Dict[str, Any], max_chars: int = 8000) -> str:
    """Serialise a parsed report to a compact text string for GPT context."""
    if parsed["type"] == "json":
        txt = f"Report type: JSON\nTop-level keys: {', '.join(parsed['top_level_keys'])}\n\n"
        txt += "Data preview (flattened):\n"
        for k, v in list(parsed["flat_preview"].items())[:80]:
            txt += f"  {k}: {v}\n"
    elif parsed["type"] == "excel":
        txt = f"Report type: Excel Workbook\nSheets: {len(parsed['sheets'])}\n\n"
        for sheet in parsed["sheets"]:
            txt += f"=== Sheet: {sheet['sheet_name']} ===\n"
            txt += f"Rows: {sheet['row_count']}  |  Columns: {sheet['column_count']}\n"
            txt += f"Columns: {', '.join(sheet['columns'])}\n"
            if sheet["kpi_columns"]:
                txt += f"Numeric KPI columns: {', '.join(sheet['kpi_columns'])}\n"
                for col in sheet["kpi_columns"]:
                    s = sheet["column_stats"].get(col, {})
                    txt += (f"  {col}: min={s.get('min')}, max={s.get('max')}, "
                            f"mean={s.get('mean')}, sum={s.get('sum')}\n")
            if sheet["sample_rows"]:
                txt += "Sample rows:\n"
                for row in sheet["sample_rows"]:
                    txt += "  " + ", ".join(f"{k}={v}" for k, v in row.items()) + "\n"
            txt += "\n"
    else:
        txt = f"Report type: Power BI (.pbix)\n"
        txt += f"Files inside archive: {', '.join(parsed['files_found'][:20])}\n\n"
        for section, content in parsed["sections"].items():
            txt += f"\n=== {section.replace('_',' ').title()} ===\n"
            txt += json.dumps(content, indent=2)[:2000] + "\n"

    return txt[:max_chars]


# ─────────────────────────────────────────────────────────────────
# In-memory store (single-process; replaced on each upload)
# ─────────────────────────────────────────────────────────────────
_report_store: Dict[str, Any] = {}


# ─────────────────────────────────────────────────────────────────
# Endpoints
# ─────────────────────────────────────────────────────────────────

@router.post("/upload")
async def upload_report(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Accept a .pbix or .json file, parse it, generate an AI summary,
    and return the structured data + summary.
    """
    filename  = file.filename or "upload"
    raw       = await file.read()
    ext       = filename.rsplit(".", 1)[-1].lower()

    if ext not in ("pbix", "json", "xlsx", "xls"):
        raise HTTPException(
            status_code=400,
            detail="Only .pbix, .json, .xlsx, and .xls files are supported."
        )

    # Parse
    if ext == "pbix":
        parsed = _parse_pbix(raw, filename)
    elif ext in ("xlsx", "xls"):
        parsed = _parse_excel(raw, filename)
    else:
        parsed = _parse_json_report(raw)

    # Build text context for GPT
    report_text = _report_to_text(parsed)

    # Summarise with OpenAI — tailored prompt for Excel files
    client = _openai_client()
    if ext in ("xlsx", "xls"):
        system_prompt = (
            "You are an expert business intelligence analyst. "
            "The user has uploaded an Excel workbook. Your job is to: "
            "1) Identify what business domain this data covers (sales, finance, HR, operations, etc.), "
            "2) List the key KPIs and metrics found with their values (min, max, mean, total), "
            "3) Highlight notable trends, outliers, or anomalies visible in the numbers, "
            "4) Describe what a Power BI dashboard built from this data should contain "
            "   (suggested charts, slicers, and report pages), "
            "5) Give the top 3 specific, data-driven recommendations for the business. "
            "Be quantitative — cite actual numbers from the data. "
            "Format your response with clear markdown headers."
        )
    else:
        system_prompt = (
            "You are an expert business analyst. "
            "The user has uploaded a report file. "
            "Analyse the content and produce a concise executive summary covering: "
            "1) What the report is about, 2) Key metrics or KPIs found, "
            "3) Notable trends, patterns, or anomalies, "
            "4) Top 3 actionable recommendations. "
            "Be specific and quantitative where data is available. "
            "Format with clear sections using markdown headers."
        )
    summary = _chat(client, system_prompt, f"Report content:\n\n{report_text}", max_tokens=1600)

    # Store for follow-up chat
    report_id = f"{ext}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    _report_store[report_id] = {
        "filename": filename,
        "parsed":   parsed,
        "text":     report_text,
        "summary":  summary,
        "uploaded_at": datetime.now().isoformat(),
    }

    return {
        "report_id":   report_id,
        "filename":    filename,
        "type":        ext,
        "summary":     summary,
        "parsed":      parsed,
        "uploaded_at": _report_store[report_id]["uploaded_at"],
    }


@router.get("/list")
async def list_reports() -> Dict[str, Any]:
    """Return metadata for all stored reports."""
    return {
        "reports": [
            {
                "report_id":   rid,
                "filename":    v["filename"],
                "type":        v["parsed"]["type"],
                "uploaded_at": v["uploaded_at"],
            }
            for rid, v in _report_store.items()
        ]
    }


@router.get("/{report_id}")
async def get_report(report_id: str) -> Dict[str, Any]:
    """Return the full parsed data + summary for a stored report."""
    if report_id not in _report_store:
        raise HTTPException(status_code=404, detail="Report not found.")
    rec = _report_store[report_id]
    return {
        "report_id":   report_id,
        "filename":    rec["filename"],
        "type":        rec["parsed"]["type"],
        "summary":     rec["summary"],
        "parsed":      rec["parsed"],
        "uploaded_at": rec["uploaded_at"],
    }


class ReportChatRequest(BaseModel):
    report_id: str
    message: str
    history: Optional[List[Dict[str, str]]] = None   # [{role, content}, ...]


@router.post("/chat")
async def report_chat(req: ReportChatRequest) -> Dict[str, Any]:
    """
    Ask questions about a previously uploaded report.
    Maintains a short history for multi-turn conversation.
    """
    if req.report_id not in _report_store:
        raise HTTPException(status_code=404, detail="Report not found. Please upload it first.")

    rec = _report_store[req.report_id]
    client = _openai_client()

    system_prompt = (
        "You are an expert business analyst assistant. "
        "The user is asking questions about the following report. "
        "Answer concisely and precisely, citing specific data from the report where available. "
        f"\n\n=== REPORT: {rec['filename']} ===\n{rec['text']}\n\n"
        f"=== EXECUTIVE SUMMARY ===\n{rec['summary']}"
    )

    # Build messages list with optional history
    messages = [{"role": "system", "content": system_prompt}]
    if req.history:
        for h in req.history[-6:]:   # keep last 6 turns to stay within token budget
            messages.append({"role": h["role"], "content": h["content"]})
    messages.append({"role": "user", "content": req.message})

    model = settings.OPENAI_MODEL or "gpt-4o-mini"
    try:
        rsp = client.chat.completions.create(
            model=model,
            messages=messages,
            max_tokens=900,
            temperature=0.4,
        )
        answer = rsp.choices[0].message.content.strip()
    except Exception:
        try:
            rsp = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=messages,
                max_tokens=900,
                temperature=0.4,
            )
            answer = rsp.choices[0].message.content.strip()
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"OpenAI error: {e}")

    return {
        "answer":    answer,
        "report_id": req.report_id,
        "timestamp": datetime.now().isoformat(),
    }


@router.delete("/{report_id}")
async def delete_report(report_id: str) -> Dict[str, Any]:
    """Remove a report from the in-memory store."""
    if report_id not in _report_store:
        raise HTTPException(status_code=404, detail="Report not found.")
    del _report_store[report_id]
    return {"deleted": report_id}

# Made with Bob
